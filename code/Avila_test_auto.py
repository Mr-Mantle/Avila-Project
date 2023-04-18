from Py_torch.Avila_project.Avila_Project import *
import openpyxl

if __name__ == '__main__':
    def auto_test_AESE(test_num=5):
        path = "Avila_test_result.xlsx"
        wb = openpyxl.load_workbook(path)
        sheet = wb['AESENet']
        for index in (np.arange(5 * test_num) + 1):
            if index % 5 == 0:
                sheet.cell(row=index - 4, column=1, value=f"Test {int(index / 5)}")
                sheet.cell(row=index - 3, column=1, value="Weighted recon loss:")
                sheet.cell(row=index - 2, column=1, value="True recon loss:")
                sheet.cell(row=index - 1, column=1, value="sorted features:")
                sheet.cell(row=index, column=1, value="sorted weights(%):")
                dataset = AvilaDataset(split=True)
                dataset.generate(to_tensor=True)
                model = AvilaAESEDoubleLoss(dataset.n_input, h=9)
                tr = TrainingProcess(model, dataset, double_loss=True, weights=(6, 1))
                tr.training(epochs=1000)
                test_acc = tr.evaluation(record=True)
                sorted_features, sorted_weights_pc = tr.get_SE_weight()
                weighted_recon_loss, true_recon_loss = tr.plot_dl_loss(record=True)
                sheet.cell(row=index - 4, column=2, value=test_acc)
                sheet.cell(row=index - 3, column=2, value=weighted_recon_loss)
                sheet.cell(row=index - 2, column=2, value=true_recon_loss)
                for col_idx, item in enumerate(sorted_features, start=2):
                    sheet.cell(row=index - 1, column=col_idx, value=item)
                for col_idx, item in enumerate(sorted_weights_pc, start=2):
                    sheet.cell(row=index, column=col_idx, value=item)
            wb.save(path)


    def auto_test_AE(test_num=5):
        path = "Avila_test_result.xlsx"
        wb = openpyxl.load_workbook(path)
        sheet = wb['AENet']
        sheet.cell(row=1, column=1, value="Test number")
        sheet.cell(row=1, column=2, value="Accuracy")
        sheet.cell(row=1, column=3, value="Weighted recon loss")
        sheet.cell(row=1, column=4, value="True recon loss")
        for index in (np.arange(test_num) + 2):
            sheet.cell(row=index, column=1, value=f"Test{index - 1}")
            dataset = AvilaDataset(split=True)
            dataset.generate(to_tensor=True)
            model = AvilaAEDoubleLoss(dataset.n_input, h=9)
            tr = TrainingProcess(model, dataset, double_loss=True, weights=(6, 1))
            tr.training(epochs=1000)
            test_acc = tr.evaluation(record=True)
            weighted_recon_loss, true_recon_loss = tr.plot_dl_loss(record=True)
            sheet.cell(row=index, column=2, value=test_acc)
            sheet.cell(row=index, column=3, value=weighted_recon_loss)
            sheet.cell(row=index, column=4, value=true_recon_loss)
            wb.save(path)

    def auto_test_AE_hidden(test_num=9):
        path = "Avila_test_result.xlsx"
        wb = openpyxl.load_workbook(path)
        sheet = wb['AE_PCA_SVD']
        for index in (np.arange(test_num) + 1):
            dataset = AvilaDataset(split=True)
            dataset.generate(to_tensor=True)
            model = AvilaAEDoubleLoss(dataset.n_input, h=index)
            tr = TrainingProcess(model, dataset, double_loss=True, weights=(6, 1))
            tr.training(epochs=1000)
            test_acc = tr.evaluation(record=True)
            sheet.cell(row=index + 1, column=4, value=test_acc)
            wb.save(path)


    auto_test_AE_hidden()




